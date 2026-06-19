"""
Tracer Metrics Analysis
=======================

Python routine for calculating sediment-tracer recovery, centre-of-mass
displacement, spreading-based diffusion, transport velocity, and transport-flux metrics from a
multi-campaign tracer-monitoring workbook.

The workflow reads an Excel workbook with one injection/reference sheet and one
or more campaign sheets, projects Cartesian x/y offsets into along- and
cross-sand-body components, calculates tracer recovery and transport metrics,
and writes a structured Excel output workbook.

Expected workbook structure
---------------------------
Sheets:
- IP: reference injection point / initial condition
- C1, C2, ...: sampling campaigns

Minimum required columns in campaign sheets:
- SAMPLE_ID
- DATE_TIME
- x
- y
- Ac
- m
- Tp_area

Optional columns:
- tracer_nr
- photo_area
- dreger_d
- sed_poro
- rho_sed
- tracer_m
- delta_mix
"""

from __future__ import annotations

import argparse
import math
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


INPUT_PREFIX = "INPUT_"
OUTPUT_PREFIX = "OUTPUT_"

DEFAULT_INPUT_FILE = "INPUT_TRACER_DATA.xlsx"
DEFAULT_OUTPUT_SUFFIX = "TRACER_METRICS"

OUTPUT_ALONG = "asb"
OUTPUT_CROSS = "csb"
ALONG_AZIMUTH_DEG = 41.0
CROSS_AZIMUTH_DEG = 131.0

GLOBAL_ORIGIN_MODE = "ip_relative_projected_offsets"
TARGET_SHEETS = [f"C{i}" for i in range(1, 6)]


def safe_tag(text: str) -> str:
    """Return a filesystem-safe label."""
    return str(text).strip().replace(" ", "_").replace("/", "_").replace("\\", "_")


def default_output_path(input_path: Path) -> Path:
    """Build a clear output workbook path from the input workbook name."""
    output_stem = input_path.stem
    if output_stem.startswith(INPUT_PREFIX):
        output_stem = output_stem.replace(INPUT_PREFIX, OUTPUT_PREFIX, 1)
    else:
        output_stem = f"{OUTPUT_PREFIX}{output_stem}"
    output_name = f"{output_stem}_{DEFAULT_OUTPUT_SUFFIX}.xlsx"
    return input_path.with_name(output_name)


def scaled_1e8(value: float) -> float:
    """Scale a value by 1e8 and round to two decimals."""
    if pd.isna(value):
        return float("nan")
    return round(float(value) * 1e8, 2)


def parse_datetime_series(series: pd.Series) -> pd.Series:
    """Parse campaign date/time values."""
    values = series.astype(str).str.strip()
    return pd.to_datetime(values, errors="coerce", dayfirst=False)


def project_xy_to_sandbody(dx, dy):
    """
    Project Cartesian east/north offsets into along-/cross-sand-body components.

    Component along azimuth A is:
    component = dx * sin(A) + dy * cos(A)
    """
    along_rad = math.radians(ALONG_AZIMUTH_DEG)
    cross_rad = math.radians(CROSS_AZIMUTH_DEG)
    along_component = dx * math.sin(along_rad) + dy * math.cos(along_rad)
    cross_component = dx * math.sin(cross_rad) + dy * math.cos(cross_rad)
    return along_component, cross_component


def rotate_vector_xy_to_sandbody(vx: float, vy: float):
    """Project a Cartesian vector into along-/cross-sand-body components."""
    if pd.isna(vx) or pd.isna(vy):
        return float("nan"), float("nan")
    return project_xy_to_sandbody(vx, vy)


def compute_as_cs_from_xy(x: pd.Series, y: pd.Series, origin_x: float, origin_y: float):
    """Compute diagnostic sand-body coordinates relative to a local origin."""
    x_rel = x.astype(float) - origin_x
    y_rel = y.astype(float) - origin_y
    return project_xy_to_sandbody(x_rel, y_rel)


def standardize_campaign_df(df: pd.DataFrame, sheet_name: str) -> pd.DataFrame:
    """
    Rename workbook-specific columns to internal names and validate coordinates.

    Existing asb/csb columns are preserved only as input diagnostics. The routine
    recomputes local projected components from x/y offsets.
    """
    rename_map = {
        "SAMPLE_ID": "sample",
        "DATE_TIME": "End_date",
    }
    out = df.rename(columns=rename_map).copy()

    if "Ac " in out.columns and "Ac" not in out.columns:
        out = out.rename(columns={"Ac ": "Ac"})

    missing_xy = [col for col in ["sample", "x", "y"] if col not in out.columns]
    if missing_xy:
        raise KeyError(f"Missing required columns {missing_xy} in sheet '{sheet_name}'.")

    if "End_date" not in out.columns:
        raise KeyError(f"Missing DATE_TIME/End_date column in sheet '{sheet_name}'.")

    out["End_date"] = parse_datetime_series(out["End_date"])
    out["x"] = pd.to_numeric(out["x"], errors="coerce")
    out["y"] = pd.to_numeric(out["y"], errors="coerce")
    out["as"] = float("nan")
    out["cs"] = float("nan")

    out = out.dropna(subset=["x", "y"])
    if out.empty:
        raise ValueError(f"No valid coordinate records found in sheet '{sheet_name}'.")

    return out


def restore_output_column_names(df: pd.DataFrame) -> pd.DataFrame:
    """Rename internal along/cross columns to selected output labels."""
    rename_map = {
        "as": OUTPUT_ALONG,
        "cs": OUTPUT_CROSS,
        "Dist_as": f"Dist_{OUTPUT_ALONG}",
        "Dist_cs": f"Dist_{OUTPUT_CROSS}",
        "TMcal*Dist_as": f"TMcal*Dist_{OUTPUT_ALONG}",
        "TMcal*Dist_cs": f"TMcal*Dist_{OUTPUT_CROSS}",
        "Dev_as": f"Dev_{OUTPUT_ALONG}",
        "Dev_cs": f"Dev_{OUTPUT_CROSS}",
        "TMcal*Dev_as^2": f"TMcal*Dev_{OUTPUT_ALONG}^2",
        "TMcal*Dev_cs^2": f"TMcal*Dev_{OUTPUT_CROSS}^2",
        "TMcal*as": f"TMcal*{OUTPUT_ALONG}",
        "TMcal*cs": f"TMcal*{OUTPUT_CROSS}",
        "CM_as": f"CM_{OUTPUT_ALONG}",
        "CM_cs": f"CM_{OUTPUT_CROSS}",
    }
    return df.rename(columns={k: v for k, v in rename_map.items() if k in df.columns})


def get_first_valid_datetime(df: pd.DataFrame, sheet_name: str) -> pd.Timestamp:
    """Return the first valid datetime for a campaign sheet."""
    if "End_date" not in df.columns:
        raise KeyError(f"'End_date' not found in sheet '{sheet_name}'.")

    valid_dates = df["End_date"].dropna()
    if valid_dates.empty:
        raise ValueError(f"Could not parse DATE_TIME/End_date values in sheet '{sheet_name}'.")
    return valid_dates.iloc[0]


def build_reference_from_ip(ip_df: pd.DataFrame) -> dict:
    """Build the initial reference values from the IP sheet."""
    dt0 = get_first_valid_datetime(ip_df, "IP")
    row0 = ip_df.iloc[0]

    ref = {
        "previous_end_date": dt0,
        "prev_CMx": float(row0["x"]),
        "prev_CMy": float(row0["y"]),
        "prev_CMas": 0.0,
        "prev_CMcs": 0.0,
        "prev_var_x": 0.0,
        "prev_var_y": 0.0,
        "prev_var_as": 0.0,
        "prev_var_cs": 0.0,
        "origin_x": float(row0["x"]),
        "origin_y": float(row0["y"]),
        "dreger_d_global": 0.15,
        "sed_poro_global": 0.6,
        "rho_sed_global": float("nan"),
        "tracer_m_total": 100.0,
    }

    for source_col, ref_key in [
        ("dreger_d", "dreger_d_global"),
        ("sed_poro", "sed_poro_global"),
        ("rho_sed", "rho_sed_global"),
        ("tracer_m", "tracer_m_total"),
    ]:
        if source_col in ip_df.columns and not pd.isna(ip_df[source_col].iloc[0]):
            ref[ref_key] = float(ip_df[source_col].iloc[0])

    return ref


def process_campaign_sheet(
    df: pd.DataFrame,
    sheet_name: str,
    previous_end_date,
    prev_CMx,
    prev_CMy,
    prev_CMas,
    prev_CMcs,
    prev_var_x,
    prev_var_y,
    prev_var_as,
    prev_var_cs,
    as_origin_x,
    as_origin_y,
    dreger_d_global,
    sed_poro_global,
    rho_sed_global,
    tracer_m_total,
):
    """Process one sampling campaign and return processed data plus summaries."""
    this_end_date = get_first_valid_datetime(df, sheet_name)
    delta_t_seconds = (this_end_date - previous_end_date).total_seconds()
    df["delta_t"] = delta_t_seconds

    if "tracer_nr" in df.columns and "photo_area" in df.columns:
        df["tracer_grains_m2"] = df["tracer_nr"] / df["photo_area"]
    else:
        print(f"WARNING: 'tracer_nr' or 'photo_area' missing in '{sheet_name}'.")
        df["tracer_grains_m2"] = float("nan")

    for required in ["Ac", "m", "Tp_area"]:
        if required not in df.columns:
            raise KeyError(f"Missing '{required}' column in sheet '{sheet_name}'.")

    df["TMcal"] = df["Ac"] / df["m"]

    dreger_d = (
        float(df["dreger_d"].iloc[0])
        if "dreger_d" in df.columns and not pd.isna(df["dreger_d"].iloc[0])
        else dreger_d_global
    )
    sed_poro = (
        float(df["sed_poro"].iloc[0])
        if "sed_poro" in df.columns and not pd.isna(df["sed_poro"].iloc[0])
        else sed_poro_global
    )

    if "rho_sed" in df.columns and not pd.isna(df["rho_sed"].iloc[0]):
        rho_sed = float(df["rho_sed"].iloc[0])
    else:
        if math.isnan(rho_sed_global):
            raise KeyError(f"'rho_sed' not found in sheet '{sheet_name}' or in 'IP'.")
        rho_sed = rho_sed_global

    tracer_m = (
        float(df["tracer_m"].iloc[0])
        if "tracer_m" in df.columns and not pd.isna(df["tracer_m"].iloc[0])
        else tracer_m_total
    )

    solid_fraction = 1.0 - sed_poro
    if solid_fraction < 0 or solid_fraction > 1:
        raise ValueError(
            f"Invalid sed_poro value in sheet '{sheet_name}': {sed_poro}. "
            "Porosity must be between 0 and 1."
        )

    df["T_rate"] = df["TMcal"] * df["Tp_area"] * dreger_d * rho_sed * solid_fraction
    df["T_rate_%"] = 100.0 * df["T_rate"] / tracer_m

    df["Dist_x"] = df["x"] - prev_CMx
    df["Dist_y"] = df["y"] - prev_CMy
    df["Dist_as"], df["Dist_cs"] = project_xy_to_sandbody(df["Dist_x"], df["Dist_y"])
    df["as"], df["cs"] = compute_as_cs_from_xy(df["x"], df["y"], as_origin_x, as_origin_y)

    df["TMcal*Dist_x"] = df["TMcal"] * df["Dist_x"]
    df["TMcal*Dist_y"] = df["TMcal"] * df["Dist_y"]
    df["TMcal*Dist_as"] = df["TMcal"] * df["Dist_as"]
    df["TMcal*Dist_cs"] = df["TMcal"] * df["Dist_cs"]

    df["TMcal*x"] = df["TMcal"] * df["x"]
    df["TMcal*y"] = df["TMcal"] * df["y"]
    df["TMcal*as"] = df["TMcal"] * df["as"]
    df["TMcal*cs"] = df["TMcal"] * df["cs"]

    for spread_col in [
        "Dev_x",
        "Dev_y",
        "Dev_as",
        "Dev_cs",
        "TMcal*Dev_x^2",
        "TMcal*Dev_y^2",
        "TMcal*Dev_as^2",
        "TMcal*Dev_cs^2",
    ]:
        df[spread_col] = float("nan")

    tm_sum = df["TMcal"].sum()
    tm_sq_sum = (df["TMcal"] ** 2).sum()

    D_x = D_y = D_as = D_cs = float("nan")
    D_total_xy = D_total_as_cs = float("nan")
    Dr_xy = Dr_as_cs = float("nan")
    Vm_x = Vm_y = Vm_as = Vm_cs = float("nan")
    Q_x = Q_y = Q_as = Q_cs = float("nan")
    delta_mix_val = float("nan")
    grains_m2_mean = float("nan")
    T_rate_sum = float("nan")
    T_rate_pct_sum = float("nan")
    var_x = var_y = var_as = var_cs = float("nan")
    delta_var_x = delta_var_y = delta_var_as = delta_var_cs = float("nan")

    if tm_sum == 0 or math.isclose(tm_sum, 0.0):
        print(f"WARNING: Sum(TMcal) = 0 in '{sheet_name}'. Metrics set to NaN.")
        CMx = CMy = CMas = CMcs = float("nan")
    else:
        CMx = df["TMcal*x"].sum() / tm_sum
        CMy = df["TMcal*y"].sum() / tm_sum
        CMas, CMcs = project_xy_to_sandbody(CMx - as_origin_x, CMy - as_origin_y)

        grains_m2_mean = df["tracer_grains_m2"].mean()
        T_rate_sum = df["T_rate"].sum()
        T_rate_pct_sum = 100.0 * T_rate_sum / tracer_m

        df["Dev_x"] = df["x"] - CMx
        df["Dev_y"] = df["y"] - CMy
        df["Dev_as"] = df["as"] - CMas
        df["Dev_cs"] = df["cs"] - CMcs
        df["TMcal*Dev_x^2"] = df["TMcal"] * df["Dev_x"] ** 2
        df["TMcal*Dev_y^2"] = df["TMcal"] * df["Dev_y"] ** 2
        df["TMcal*Dev_as^2"] = df["TMcal"] * df["Dev_as"] ** 2
        df["TMcal*Dev_cs^2"] = df["TMcal"] * df["Dev_cs"] ** 2

        var_x = df["TMcal*Dev_x^2"].sum() / tm_sum
        var_y = df["TMcal*Dev_y^2"].sum() / tm_sum
        var_as = df["TMcal*Dev_as^2"].sum() / tm_sum
        var_cs = df["TMcal*Dev_cs^2"].sum() / tm_sum
        delta_var_x = var_x - prev_var_x
        delta_var_y = var_y - prev_var_y
        delta_var_as = var_as - prev_var_as
        delta_var_cs = var_cs - prev_var_cs

        if delta_t_seconds == 0:
            print(f"WARNING: zero delta_t in '{sheet_name}'. Diffusion set to NaN.")
        else:
            # Miller & Komar / method-of-moments style spreading coefficient:
            # K = 0.5 * d(sigma^2) / dt, using TMcal-weighted variances.
            denom = 2.0 * delta_t_seconds
            D_x = delta_var_x / denom
            D_y = delta_var_y / denom
            D_as = delta_var_as / denom
            D_cs = delta_var_cs / denom
            D_total_xy = D_x + D_y
            D_total_as_cs = D_as + D_cs
            Dr_xy = math.sqrt(D_x ** 2 + D_y ** 2)
            Dr_as_cs = math.sqrt(D_as ** 2 + D_cs ** 2)

        dist_cm_x = CMx - prev_CMx
        dist_cm_y = CMy - prev_CMy

        if delta_t_seconds != 0:
            Vm_x = dist_cm_x / delta_t_seconds
            Vm_y = dist_cm_y / delta_t_seconds
            Vm_as, Vm_cs = rotate_vector_xy_to_sandbody(Vm_x, Vm_y)
        else:
            print(f"WARNING: delta_t = 0 in '{sheet_name}'. Velocities set to NaN.")

        delta_mix_col = next((col for col in ["delta_mix", "delta_mix "] if col in df.columns), None)
        if delta_mix_col is not None:
            delta_mix_series = df[delta_mix_col].dropna()
            if not delta_mix_series.empty:
                delta_mix_val = float(delta_mix_series.iloc[0])
                Q_x = Vm_x * delta_mix_val
                Q_y = Vm_y * delta_mix_val
                Q_as = Vm_as * delta_mix_val
                Q_cs = Vm_cs * delta_mix_val
            else:
                print(f"WARNING: 'delta_mix' column empty in '{sheet_name}'.")
        else:
            print(f"WARNING: no 'delta_mix' column found in '{sheet_name}'.")

    dist_cm_x = CMx - prev_CMx
    dist_cm_y = CMy - prev_CMy
    dist_cm_as, dist_cm_cs = project_xy_to_sandbody(dist_cm_x, dist_cm_y)

    cm_sheet_name = "CM" + sheet_name[1:]
    cm_df = pd.DataFrame({
        "CM": [cm_sheet_name],
        "CM_x": [CMx],
        "CM_y": [CMy],
        f"CM_{OUTPUT_ALONG}": [CMas],
        f"CM_{OUTPUT_CROSS}": [CMcs],
        "TMcal_sum": [tm_sum],
        "TMcal_sq_sum": [tm_sq_sum],
        "delta_t_s": [delta_t_seconds],
        "Var_x": [var_x],
        "Var_y": [var_y],
        f"Var_{OUTPUT_ALONG}": [var_as],
        f"Var_{OUTPUT_CROSS}": [var_cs],
        "Prev_Var_x": [prev_var_x],
        "Prev_Var_y": [prev_var_y],
        f"Prev_Var_{OUTPUT_ALONG}": [prev_var_as],
        f"Prev_Var_{OUTPUT_CROSS}": [prev_var_cs],
        "Delta_Var_x": [delta_var_x],
        "Delta_Var_y": [delta_var_y],
        f"Delta_Var_{OUTPUT_ALONG}": [delta_var_as],
        f"Delta_Var_{OUTPUT_CROSS}": [delta_var_cs],
        "DistCM_x": [dist_cm_x],
        "DistCM_y": [dist_cm_y],
        f"DistCM_{OUTPUT_ALONG}": [dist_cm_as],
        f"DistCM_{OUTPUT_CROSS}": [dist_cm_cs],
        "TCOMd_xy": [math.sqrt(dist_cm_x ** 2 + dist_cm_y ** 2)],
        f"TCOMd_{OUTPUT_ALONG}{OUTPUT_CROSS}": [math.sqrt(dist_cm_as ** 2 + dist_cm_cs ** 2)],
        "TCOMdir_deg_xy": [(math.degrees(math.atan2(dist_cm_x, dist_cm_y)) + 360) % 360],
        f"TCOMdir_deg_{OUTPUT_ALONG}{OUTPUT_CROSS}": [
            (math.degrees(math.atan2(dist_cm_as, dist_cm_cs)) + 360) % 360
        ],
        "D_x": [D_x],
        "D_y": [D_y],
        f"D_{OUTPUT_ALONG}": [D_as],
        f"D_{OUTPUT_CROSS}": [D_cs],
        "D_total_xy": [D_total_xy],
        f"D_total_{OUTPUT_ALONG}_{OUTPUT_CROSS}": [D_total_as_cs],
        "Dr_xy": [Dr_xy],
        f"Dr_{OUTPUT_ALONG}_{OUTPUT_CROSS}": [Dr_as_cs],
        "Vm_x": [Vm_x],
        "Vm_y": [Vm_y],
        f"Vm_{OUTPUT_ALONG}": [Vm_as],
        f"Vm_{OUTPUT_CROSS}": [Vm_cs],
        "delta_mix": [delta_mix_val],
        "Q_x": [Q_x],
        "Q_y": [Q_y],
        f"Q_{OUTPUT_ALONG}": [Q_as],
        f"Q_{OUTPUT_CROSS}": [Q_cs],
        "tracer_grains_m2_mean": [grains_m2_mean],
        "T_rate_sum": [T_rate_sum],
        "T_rate_%_sum": [T_rate_pct_sum],
        "prev_EndDate": [previous_end_date],
        "End_date": [this_end_date],
        "origin_x": [as_origin_x],
        "origin_y": [as_origin_y],
        "asb_azimuth_deg": [ALONG_AZIMUTH_DEG],
        "csb_azimuth_deg": [CROSS_AZIMUTH_DEG],
    })

    return (
        restore_output_column_names(df),
        cm_sheet_name,
        cm_df,
        this_end_date,
        CMx,
        CMy,
        CMas,
        CMcs,
        var_x,
        var_y,
        var_as,
        var_cs,
    )


def metadata_table(input_file: Path, output_file: Path, campaign_sheets: list[str]) -> pd.DataFrame:
    """Build metadata for the output workbook."""
    records = [
        ("routine", "tracer_metrics_analysis"),
        ("input_file", input_file.name),
        ("output_file", output_file.name),
        ("origin_mode", GLOBAL_ORIGIN_MODE),
        ("origin_x", "IP/reference x coordinate; see CM sheets"),
        ("origin_y", "IP/reference y coordinate; see CM sheets"),
        ("transport_distance_mode", "sample/CM displacement from previous campaign CM or IP"),
        ("diffusion_mode", "method of moments; time-rate of change of weighted variance"),
        ("diffusion_convention", "D = delta_variance / (2 * delta_t), using TMcal weights"),
        ("first_interval_variance_assumption", "IP/reference variance = 0"),
        ("asb_azimuth_deg", ALONG_AZIMUTH_DEG),
        ("csb_azimuth_deg", CROSS_AZIMUTH_DEG),
        ("campaign_sheets_expected", ", ".join(campaign_sheets)),
        ("output_along_label", OUTPUT_ALONG),
        ("output_cross_label", OUTPUT_CROSS),
    ]
    return pd.DataFrame(records, columns=["parameter", "value"])


def save_output_workbook(
    output_file: Path,
    input_file: Path,
    ip_df: pd.DataFrame,
    processed_sheets: dict,
    com_sheets: dict,
    vmq_df: pd.DataFrame,
    diff_df: pd.DataFrame,
    recovery_df: pd.DataFrame,
    campaign_sheets: list[str],
) -> None:
    """Save all processed outputs to an Excel workbook."""
    output_file.parent.mkdir(parents=True, exist_ok=True)
    metadata_df = metadata_table(input_file, output_file, campaign_sheets)

    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        restore_output_column_names(ip_df).to_excel(writer, sheet_name="IP", index=False)
        for name, df_out in processed_sheets.items():
            df_out.to_excel(writer, sheet_name=name, index=False)
        for cm_name, cm_df in com_sheets.items():
            cm_df.to_excel(writer, sheet_name=cm_name, index=False)
        vmq_df.to_excel(writer, sheet_name="Velocity_Flux_1e-8", index=False)
        diff_df.to_excel(writer, sheet_name="Diffusion_1e-8", index=False)
        recovery_df.to_excel(writer, sheet_name="Tracer_Recovery", index=False)
        metadata_df.to_excel(writer, sheet_name="Metadata", index=False)

    workbook = load_workbook(output_file)
    for sheet_name in ["Velocity_Flux_1e-8", "Diffusion_1e-8", "Tracer_Recovery"]:
        if sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            for row in worksheet.iter_rows(min_row=2, min_col=2):
                for cell in row:
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = "0.00"
    workbook.save(output_file)


def process_workbook(input_file: Path, output_file: Path, campaign_sheets: list[str]) -> None:
    """Run the complete tracer-metrics workflow."""
    if not input_file.exists():
        raise FileNotFoundError(f"Input workbook not found: {input_file}")

    xls = pd.ExcelFile(input_file)
    if "IP" not in xls.sheet_names:
        raise ValueError("Sheet 'IP' not found. It is required as reference.")

    ip_raw = pd.read_excel(xls, sheet_name="IP")
    ip_df = standardize_campaign_df(ip_raw, "IP")
    ref = build_reference_from_ip(ip_df)

    previous_end_date = ref["previous_end_date"]
    prev_CMx = ref["prev_CMx"]
    prev_CMy = ref["prev_CMy"]
    prev_CMas = ref["prev_CMas"]
    prev_CMcs = ref["prev_CMcs"]
    prev_var_x = ref["prev_var_x"]
    prev_var_y = ref["prev_var_y"]
    prev_var_as = ref["prev_var_as"]
    prev_var_cs = ref["prev_var_cs"]
    as_origin_x = ref["origin_x"]
    as_origin_y = ref["origin_y"]

    ip_df["as"], ip_df["cs"] = compute_as_cs_from_xy(ip_df["x"], ip_df["y"], as_origin_x, as_origin_y)

    processed_sheets = {}
    com_sheets = {}
    vmq_rows = []
    diff_rows = []
    prev_label = "IP"

    for sheet_name in campaign_sheets:
        if sheet_name not in xls.sheet_names:
            print(f"WARNING: sheet '{sheet_name}' not found in workbook. Skipping.")
            continue

        print(f"Processing sheet: {sheet_name}")
        df_raw = pd.read_excel(xls, sheet_name=sheet_name)
        df = standardize_campaign_df(df_raw, sheet_name)

        (
            processed_df,
            cm_sheet_name,
            cm_df,
            this_end_date,
            CMx,
            CMy,
            CMas,
            CMcs,
            var_x,
            var_y,
            var_as,
            var_cs,
        ) = process_campaign_sheet(
            df=df,
            sheet_name=sheet_name,
            previous_end_date=previous_end_date,
            prev_CMx=prev_CMx,
            prev_CMy=prev_CMy,
            prev_CMas=prev_CMas,
            prev_CMcs=prev_CMcs,
            prev_var_x=prev_var_x,
            prev_var_y=prev_var_y,
            prev_var_as=prev_var_as,
            prev_var_cs=prev_var_cs,
            as_origin_x=as_origin_x,
            as_origin_y=as_origin_y,
            dreger_d_global=ref["dreger_d_global"],
            sed_poro_global=ref["sed_poro_global"],
            rho_sed_global=ref["rho_sed_global"],
            tracer_m_total=ref["tracer_m_total"],
        )

        processed_sheets[sheet_name] = processed_df
        com_sheets[cm_sheet_name] = cm_df

        row = cm_df.iloc[0]
        interval_label = f"{prev_label}-{sheet_name}"

        Q_x = float(row["Q_x"])
        Q_y = float(row["Q_y"])
        Q_as = float(row[f"Q_{OUTPUT_ALONG}"])
        Q_cs = float(row[f"Q_{OUTPUT_CROSS}"])

        D_x = float(row["D_x"])
        D_y = float(row["D_y"])
        D_as = float(row[f"D_{OUTPUT_ALONG}"])
        D_cs = float(row[f"D_{OUTPUT_CROSS}"])
        D_total_xy = float(row["D_total_xy"])
        D_total_as_cs = float(row[f"D_total_{OUTPUT_ALONG}_{OUTPUT_CROSS}"])
        Dr_xy = float(row["Dr_xy"])
        Dr_as_cs = float(row[f"Dr_{OUTPUT_ALONG}_{OUTPUT_CROSS}"])
        delta_var_x = float(row["Delta_Var_x"])
        delta_var_y = float(row["Delta_Var_y"])
        delta_var_as = float(row[f"Delta_Var_{OUTPUT_ALONG}"])
        delta_var_cs = float(row[f"Delta_Var_{OUTPUT_CROSS}"])

        Vm_x = float(row["Vm_x"])
        Vm_y = float(row["Vm_y"])
        Vm_as = float(row[f"Vm_{OUTPUT_ALONG}"])
        Vm_cs = float(row[f"Vm_{OUTPUT_CROSS}"])

        Q_xy_res = float("nan") if any(math.isnan(v) for v in (Q_x, Q_y)) else math.sqrt(Q_x ** 2 + Q_y ** 2)
        Q_as_cs_res = float("nan") if any(math.isnan(v) for v in (Q_as, Q_cs)) else math.sqrt(Q_as ** 2 + Q_cs ** 2)

        vmq_rows.append({
            "Sampling intervals": interval_label,
            "Vm_x": scaled_1e8(Vm_x),
            "Vm_y": scaled_1e8(Vm_y),
            f"Vm_{OUTPUT_ALONG}": scaled_1e8(Vm_as),
            f"Vm_{OUTPUT_CROSS}": scaled_1e8(Vm_cs),
            "Q_x": scaled_1e8(Q_x),
            "Q_y": scaled_1e8(Q_y),
            f"Q_{OUTPUT_ALONG}": scaled_1e8(Q_as),
            f"Q_{OUTPUT_CROSS}": scaled_1e8(Q_cs),
            "Q_xy_resultant": scaled_1e8(Q_xy_res),
            f"Q_{OUTPUT_ALONG}_{OUTPUT_CROSS}_resultant": scaled_1e8(Q_as_cs_res),
        })

        diff_rows.append({
            "Sampling intervals": interval_label,
            "Delta_Var_x": delta_var_x,
            "Delta_Var_y": delta_var_y,
            f"Delta_Var_{OUTPUT_ALONG}": delta_var_as,
            f"Delta_Var_{OUTPUT_CROSS}": delta_var_cs,
            "D_x": scaled_1e8(D_x),
            "D_y": scaled_1e8(D_y),
            f"D_{OUTPUT_ALONG}": scaled_1e8(D_as),
            f"D_{OUTPUT_CROSS}": scaled_1e8(D_cs),
            "D_total_xy": scaled_1e8(D_total_xy),
            f"D_total_{OUTPUT_ALONG}_{OUTPUT_CROSS}": scaled_1e8(D_total_as_cs),
            "Dr_xy": scaled_1e8(Dr_xy),
            f"Dr_{OUTPUT_ALONG}_{OUTPUT_CROSS}": scaled_1e8(Dr_as_cs),
        })

        previous_end_date = this_end_date
        prev_CMx = CMx
        prev_CMy = CMy
        prev_CMas = CMas
        prev_CMcs = CMcs
        prev_var_x = var_x
        prev_var_y = var_y
        prev_var_as = var_as
        prev_var_cs = var_cs
        prev_label = sheet_name

    vmq_df = pd.DataFrame(vmq_rows)
    diff_df = pd.DataFrame(diff_rows)

    recovery_rows = []
    campaign_order = sorted(com_sheets.keys(), key=lambda name: int(name[2:]))
    for cm_name in campaign_order:
        row = com_sheets[cm_name].iloc[0]
        recovery_rows.append({
            "Campaign": cm_name,
            "T_rate_sum_kg": float(row["T_rate_sum"]),
            "T_rate_%": float(row["T_rate_%_sum"]),
        })

    recovery_df = pd.DataFrame(recovery_rows)
    if not recovery_df.empty:
        recovery_df["T_rate_%_cum"] = recovery_df["T_rate_%"].cumsum()
        recovery_df.loc[len(recovery_df.index)] = {
            "Campaign": "TOTAL",
            "T_rate_sum_kg": recovery_df["T_rate_sum_kg"].sum(),
            "T_rate_%": recovery_df["T_rate_%"].sum(),
            "T_rate_%_cum": recovery_df["T_rate_%_cum"].iloc[-1],
        }

    save_output_workbook(
        output_file=output_file,
        input_file=input_file,
        ip_df=ip_df,
        processed_sheets=processed_sheets,
        com_sheets=com_sheets,
        vmq_df=vmq_df,
        diff_df=diff_df,
        recovery_df=recovery_df,
        campaign_sheets=campaign_sheets,
    )

    print("Processing completed successfully.")
    print(f"Output saved as: {output_file}")
    print(f"Output coordinate labels: {OUTPUT_ALONG} (along), {OUTPUT_CROSS} (cross)")
    print(f"Sand-body projection origin: x0={as_origin_x:.3f}, y0={as_origin_y:.3f}")


def build_parser() -> argparse.ArgumentParser:
    """Create command-line parser."""
    parser = argparse.ArgumentParser(
        description="Calculate sediment-tracer recovery, diffusion, velocity, and flux metrics."
    )
    parser.add_argument("--input", "-i", default=DEFAULT_INPUT_FILE, help="Input Excel workbook.")
    parser.add_argument("--output", "-o", default=None, help="Output Excel workbook.")
    parser.add_argument("--campaigns", default="C1,C2,C3,C4,C5", help="Comma-separated campaign sheet names.")
    parser.add_argument("--along-label", default=OUTPUT_ALONG, help="Output label for along-sand-body component.")
    parser.add_argument("--cross-label", default=OUTPUT_CROSS, help="Output label for cross-sand-body component.")
    parser.add_argument("--along-azimuth", type=float, default=ALONG_AZIMUTH_DEG, help="Along-sand-body azimuth in degrees clockwise from north.")
    parser.add_argument("--cross-azimuth", type=float, default=CROSS_AZIMUTH_DEG, help="Cross-sand-body azimuth in degrees clockwise from north.")
    return parser


def main(argv: list[str] | None = None) -> int:
    """Command-line entry point."""
    global OUTPUT_ALONG, OUTPUT_CROSS, ALONG_AZIMUTH_DEG, CROSS_AZIMUTH_DEG

    args = build_parser().parse_args(argv)
    OUTPUT_ALONG = args.along_label
    OUTPUT_CROSS = args.cross_label
    ALONG_AZIMUTH_DEG = args.along_azimuth
    CROSS_AZIMUTH_DEG = args.cross_azimuth

    input_file = Path(args.input)
    output_file = Path(args.output) if args.output else default_output_path(input_file)
    campaign_sheets = [sheet.strip() for sheet in args.campaigns.split(",") if sheet.strip()]

    process_workbook(input_file, output_file, campaign_sheets)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
